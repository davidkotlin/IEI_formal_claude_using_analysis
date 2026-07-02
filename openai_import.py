"""
openai_import.py
解析 OpenAI 的匯出檔並寫入 openai.db，與 Claude 的 cron_import.py 平行。

三種來源（節奏、來源皆不同，故獨立於 Claude pipeline）：
  1. 名單    users_import_*.xlsx  欄位 email, name  → 母體，人工維護
  2. Codex   codex-*.csv          每列自帶 date     → 逐日 token / session / message
  3. 網頁版  leaderboard-*.csv    無 date 欄，date 從「檔名」帶入 → 每人每日 tokens

匯入順序：名單(xlsx) → codex → web
  因為名單是母體，email 不在名單者一律略過（unknown），故名單要先進。

檔名日期慣例（網頁版必須）：
  檔名需含 YYYY-MM-DD，例如 leaderboard-2026-06-22.csv、web-20260622.csv。
  因為網頁版 CSV 是區間加總、本身不知道自己是哪天，日期只能從檔名取得。

CRUD 與前端：
  users 的增改刪走 openai_process 的 add_user / update_user / delete_user。
  本檔的 import_roster_xlsx 只是「用 Excel 餵那組 CRUD」的一條便道；
  未來若改用前端輸入框，直接呼叫同一組 CRUD 即可，這裡不用改。
  （Excel vs 輸入框仍待討論，兩條路共用同一批 DB 函式，先不鎖死。）
"""

import re
import csv
import logging
from pathlib import Path

from openai_process import (
    init_db,
    upsert_user_from_roster,
    import_codex_daily,
    import_web_daily,
)

# --- 設定 ---
INCOMING_DIR = Path(__file__).parent / "data" / "openai_incoming"
LOG_FORMAT = "%(asctime)s [%(levelname)s] %(message)s"
logging.basicConfig(level=logging.INFO, format=LOG_FORMAT)
logger = logging.getLogger(__name__)

DATE_RE = re.compile(r"(\d{4})[-_]?(\d{2})[-_]?(\d{2})")


def extract_date_from_filename(path: Path) -> str:
    """
    從檔名抓 YYYY-MM-DD。支援 2026-06-22 / 20260622 / 2026_06_22。
    抓不到回傳 None。
    """
    m = DATE_RE.search(path.name)
    if not m:
        return None
    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"


# ---------------------------------------------------------------------------
# 解析（純解析，不碰 DB；DB 寫入交給 openai_process）
# ---------------------------------------------------------------------------

def parse_codex_csv(path: Path) -> list:
    """把 Codex CSV 解析成 import_codex_daily 需要的列。"""
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            rows.append({
                "email": r.get("email", ""),
                "date": r.get("date", ""),
                "user_id": r.get("user_id", ""),
                "uncached_input": r.get("uncached_text_input_tokens", 0),
                "cached_input": r.get("cached_text_input_tokens", 0),
                "output": r.get("text_output_tokens", 0),
                "n_sessions": r.get("n_new_sessions_total", 0),
                "n_messages": r.get("n_user_messages_total", 0),
            })
    return rows


def parse_web_csv(path: Path, date: str) -> list:
    """
    把網頁版 leaderboard CSV 解析成 import_web_daily 需要的列。
    date 由呼叫端（從檔名）帶入。Credits / Lines of code 這方案全為 0，忽略。
    """
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            rows.append({
                "email": r.get("Email", ""),
                "date": date,
                "user_id": r.get("User ID", ""),
                "tokens": r.get("Tokens", 0),
            })
    return rows


def import_roster_xlsx(path: Path) -> dict:
    """
    讀名單 Excel（欄位 email, name）並 upsert 進 users。
    name 依名單原樣寫入（含那些 name 就是 email 的共用帳號）。
    需要 openpyxl。
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    header = None
    inserted = updated = skipped = 0
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [str(c).strip().lower() if c is not None else "" for c in row]
            continue
        record = dict(zip(header, row))
        email = record.get("email")
        name = record.get("name")
        if not email:
            skipped += 1
            continue
        result = upsert_user_from_roster(str(email), str(name) if name is not None else "")
        if result == "inserted":
            inserted += 1
        elif result == "updated":
            updated += 1
        else:
            skipped += 1
    wb.close()
    return {"roster_inserted": inserted, "roster_updated": updated, "roster_skipped": skipped}


# ---------------------------------------------------------------------------
# 對外便道：直接吃 bytes（給未來前端上傳用，對應 Claude 的 import_from_bytes）
# ---------------------------------------------------------------------------

def import_codex_from_path(path: Path) -> dict:
    return import_codex_daily(parse_codex_csv(path))


def import_web_from_path(path: Path) -> dict:
    date = extract_date_from_filename(path)
    if not date:
        raise ValueError(f"網頁版檔名沒有日期，無法判斷是哪天：{path.name}")
    return import_web_daily(parse_web_csv(path, date))


# ---------------------------------------------------------------------------
# cron / 手動掃描 incoming 資料夾
#   依檔名分派：users*.xlsx → 名單、codex*.csv → codex、其餘 *.csv → 網頁版
#   名單先處理（母體要先進）。
# ---------------------------------------------------------------------------

def _classify(path: Path) -> str:
    name = path.name.lower()
    if name.endswith((".xlsx", ".xls")) and "user" in name:
        return "roster"
    if name.endswith(".csv") and "codex" in name:
        return "codex"
    if name.endswith(".csv"):
        return "web"
    return "unknown"


def run():
    logger.info("=== openai_import 開始執行 ===")

    if not INCOMING_DIR.exists():
        logger.error(f"incoming 資料夾不存在：{INCOMING_DIR}，請先建立並放入匯出檔。")
        return

    init_db()

    files = [p for p in sorted(INCOMING_DIR.rglob("*")) if p.is_file()]
    # 名單優先，再 codex，再 web
    order = {"roster": 0, "codex": 1, "web": 2, "unknown": 9}
    files.sort(key=lambda p: order[_classify(p)])

    processed = 0
    for path in files:
        kind = _classify(path)
        if kind == "unknown":
            logger.info(f"略過無法辨識的檔案：{path.name}")
            continue

        logger.info(f"處理 [{kind}]：{path.name}")
        try:
            if kind == "roster":
                result = import_roster_xlsx(path)
                logger.info(
                    f"名單完成 - 新增 {result['roster_inserted']}、"
                    f"更新 {result['roster_updated']}、略過 {result['roster_skipped']}。"
                )
                # 名單是母體來源，不刪除，保留供對照
                processed += 1
                continue

            if kind == "codex":
                result = import_codex_from_path(path)
                logger.info(
                    f"Codex 完成 - 新增 {result['codex_inserted']} 筆、"
                    f"重複 {result['codex_skipped_dup']}、"
                    f"unknown {result['codex_skipped_unknown']}。"
                )
            else:  # web
                date = extract_date_from_filename(path)
                if not date:
                    logger.error(f"網頁版檔名無日期，略過：{path.name}")
                    continue
                result = import_web_from_path(path)
                logger.info(
                    f"網頁版 完成 ({date}) - 新增 {result['web_inserted']} 筆、"
                    f"重複 {result['web_skipped_dup']}、"
                    f"unknown {result['web_skipped_unknown']}。"
                )

            # 用量檔匯入成功後刪除（資料已進 db；與 Claude cron 一致）
            path.unlink()
            logger.info(f"已刪除：{path.name}")
            processed += 1

        except Exception as e:
            logger.error(f"處理失敗 {path.name}：{e}")

    logger.info(f"=== openai_import 執行完畢，共處理 {processed} 個檔案 ===")


if __name__ == "__main__":
    run()

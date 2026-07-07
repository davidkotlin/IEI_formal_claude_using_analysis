"""
openai_import.py（邏輯層，位於 backend/importers/）
OpenAI 匯出檔的「解析 + 寫入」邏輯，被 route（網頁上傳）與 openai_cron.py（排程）共用。
本檔不含 cron 入口 / 掃描資料夾 / logging —— 那些在專案根的 openai_cron.py。

三種來源：
  users*.xlsx      -> 名單（母體，人工維護）
  codex*.csv       -> Codex（每列自帶 date）
  leaderboard*.csv -> 網頁版（無 date 欄，date 從檔名 YYYY-MM-DD 帶入）
"""

import re
import csv
from pathlib import Path

from .openai_process import (
    upsert_user_from_roster,
    import_codex_daily,
    import_web_daily,
)

DATE_RE = re.compile(r"(\d{4})[-_]?(\d{2})[-_]?(\d{2})")


def extract_date_from_filename(path: Path) -> str:
    """從檔名抓 YYYY-MM-DD。支援 2026-06-22 / 20260622 / 2026_06_22。抓不到回 None。"""
    m = DATE_RE.search(path.name)
    if not m:
        return None
    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"


# --- 解析（純解析，不碰 DB）---

def parse_codex_csv(path: Path) -> list:
    """把 Codex CSV 解析成 import_codex_daily 需要的列。"""
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            rows.append({
                "email": r.get("email", ""),
                "date": r.get("date", ""),
                "user_id": r.get("user_id", ""),
                "uncached_input": r.get("uncached_text_input_tokens", 0),
                "cached_input": r.get("cached_text_input_tokens", 0),
                "output": r.get("text_output_tokens", 0),
                "n_sessions": r.get("n_new_sessions_total", 0),
                "n_messages": r.get("n_user_messages_total", 0),
            })
    return rows


def parse_web_csv(path: Path, date: str) -> list:
    """把網頁版 leaderboard CSV 解析成 import_web_daily 需要的列。date 由呼叫端（檔名）帶入。"""
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            rows.append({
                "email": r.get("Email", ""),
                "date": date,
                "user_id": r.get("User ID", ""),
                "tokens": r.get("Tokens", 0),
            })
    return rows


def import_roster_xlsx(path: Path) -> dict:
    """讀名單 Excel（欄位 email, name）並 upsert 進 users。需要 openpyxl。"""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    header = None
    inserted = updated = skipped = 0
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [str(c).strip().lower() if c is not None else "" for c in row]
            continue
        record = dict(zip(header, row))
        email = record.get("email")
        name = record.get("name")
        if not email:
            skipped += 1
            continue
        result = upsert_user_from_roster(str(email), str(name) if name is not None else "")
        if result == "inserted":
            inserted += 1
        elif result == "updated":
            updated += 1
        else:
            skipped += 1
    wb.close()
    return {"roster_inserted": inserted, "roster_updated": updated, "roster_skipped": skipped}


# --- 便道：吃路徑，一步到位解析＋寫入（cron 與 route 都可用）---

def import_codex_from_path(path: Path) -> dict:
    return import_codex_daily(parse_codex_csv(path))


def import_web_from_path(path: Path) -> dict:
    date = extract_date_from_filename(path)
    if not date:
        raise ValueError(f"網頁版檔名沒有日期，無法判斷是哪天：{path.name}")
    return import_web_daily(parse_web_csv(path, date))
